#!/usr/bin/env python3
"""Aplica selectores desplegables (data validation) y rangos con nombre dinámicos
sobre mad-map-data-v2.xlsx, sin tocar los datos.

Idea: cada columna referencial (linea_id, area_id, lab_id, modo_id, salida_id,
investigador_id, tema_id, sublinea_id, sublinea_a_id, sublinea_b_id) deja de
ser un campo de texto libre y se convierte en un dropdown que solo permite
elegir IDs realmente existentes en la hoja-entidad correspondiente.

El rango es dinámico: si agregas una nueva línea en 01_Lineas, el dropdown
de 02_Sublineas.linea_id la incluye automáticamente al recargar el archivo.

Cómo funciona técnicamente:
  1. Se definen "named ranges" globales por entidad (LineaIDs, AreaIDs, etc.)
     usando una fórmula OFFSET+COUNTA que crece con los datos.
  2. Cada celda referencial recibe Data Validation tipo "list" cuyo source
     es el named range correspondiente.
  3. La validación se aplica desde la primera fila de datos hasta varias
     cientas de filas hacia abajo, dejando espacio para crecer.

Idempotente: se puede correr varias veces sobre el mismo archivo sin duplicar
named ranges ni validaciones (las reemplaza).

Uso:
  python3 tools/apply_dropdowns.py
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "mad-map-data-v2.xlsx"

# Hasta dónde se aplica la validación (filas hacia abajo desde la primera
# de datos). Da espacio para crecer sin tener que volver a correr el script.
GROWTH_ROWS = 500

# ---------------------------------------------------------------------------
# Diccionario de "lookups": para cada named range, define la hoja origen,
# el nombre de la columna que se va a listar y la fila de encabezado.
# La primera fila de datos es header_row + 1.
# ---------------------------------------------------------------------------
LOOKUPS = {
    "LineaIDs":        ("01_Lineas",         "id",      4),
    "SublineaIDs":     ("02_Sublineas",      "id",      4),
    "AreaIDs":         ("03_Areas",          "código",  4),
    "ModoIDs":         ("04_Modos",          "id",      2),
    "SalidaIDs":       ("05_Salidas",        "id",      4),
    "LabIDs":          ("06_Laboratorios",   "id",      2),
    "InvestigadorIDs": ("07_Investigadores", "id",      4),
    "TemaIDs":         ("08_Temas",          "id",      4),
}

# ---------------------------------------------------------------------------
# Mapping de qué celdas (sheet + columna) se validan contra qué lookup.
# Cada tupla: (hoja destino, nombre de columna a validar, named range).
# ---------------------------------------------------------------------------
TARGETS = [
    ("02_Sublineas",           "linea_id",        "LineaIDs"),
    ("02_Sublineas",           "area_id",         "AreaIDs"),
    ("08_Temas",               "investigador_id", "InvestigadorIDs"),
    ("09_Sublinea_Tema",       "sublinea_id",     "SublineaIDs"),
    ("09_Sublinea_Tema",       "tema_id",         "TemaIDs"),
    ("10_Lab_Linea",           "lab_id",          "LabIDs"),
    ("10_Lab_Linea",           "linea_id",        "LineaIDs"),
    ("11_Lab_Salida",          "lab_id",          "LabIDs"),
    ("11_Lab_Salida",          "salida_id",       "SalidaIDs"),
    ("12_Investigador_Lab",    "investigador_id", "InvestigadorIDs"),
    ("12_Investigador_Lab",    "lab_id",          "LabIDs"),
    ("13_Investigador_Modo",   "investigador_id", "InvestigadorIDs"),
    ("13_Investigador_Modo",   "modo_id",         "ModoIDs"),
    ("14_Linea_Modo",          "linea_id",        "LineaIDs"),
    ("14_Linea_Modo",          "modo_id",         "ModoIDs"),
    ("18_Proximidad_Tematica", "sublinea_a_id",   "SublineaIDs"),
    ("18_Proximidad_Tematica", "sublinea_b_id",   "SublineaIDs"),
]


def column_letter_for_header(ws, header_row, header_name):
    """Devuelve la letra de columna ('A','B','C'...) cuya celda en `header_row`
    coincide con `header_name`. Lanza ValueError si no la encuentra.
    Usado por build_named_range y add_validation para localizar columnas
    referenciales sin asumir un orden fijo.
    """
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row, c).value == header_name:
            return ws.cell(header_row, c).column_letter
    raise ValueError(
        f"No se encontró la columna {header_name!r} en hoja {ws.title!r} "
        f"(fila de encabezados {header_row})."
    )


def build_named_range_formula(sheet_name, column_letter, first_data_row):
    """Construye la fórmula dinámica para un named range que cubre toda la
    columna de IDs desde la primera fila de datos hasta la última no vacía.

    OFFSET(start, 0, 0, COUNTA(rango_amplio), 1):
      - start: primera celda de datos (ej. A5).
      - height: número de celdas no vacías en un rango amplio que arranca
        en la misma celda (cubre crecimiento futuro).
      - width: 1 columna.

    Comillas simples alrededor del nombre de hoja para tolerar nombres con
    underscore al inicio (ej. '01_Lineas').
    """
    sheet_ref = f"'{sheet_name}'"
    start = f"{sheet_ref}!${column_letter}${first_data_row}"
    range_amplio = f"{sheet_ref}!${column_letter}${first_data_row}:${column_letter}$10000"
    return f"OFFSET({start},0,0,COUNTA({range_amplio}),1)"


def reset_defined_name(wb, name, formula):
    """Crea o reemplaza un named range a nivel del workbook. Llamado por
    apply_dropdowns para cada entrada de LOOKUPS. Idempotente: si ya
    existe, lo borra antes de re-crearlo.
    """
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(name=name, attr_text=formula)


def remove_existing_validations_for_column(ws, col_letter):
    """Elimina cualquier data validation previa que cubra esa columna en la
    hoja, para evitar acumular reglas al re-correr el script.

    Solo borra validaciones cuyo sqref sea exactamente la columna que vamos
    a re-aplicar; deja intactas otras (ej. estado=DESCARTADO en
    18_Proximidad_Tematica.G).
    """
    keep = []
    for dv in ws.data_validations.dataValidation:
        sqref = str(dv.sqref) if dv.sqref else ""
        # Match si la primera celda del sqref empieza con la columna objetivo
        # y la ÚNICA columna afectada es col_letter.
        first_cell = sqref.split()[0].split(":")[0] if sqref else ""
        first_col_letters = "".join(ch for ch in first_cell if ch.isalpha())
        if first_col_letters != col_letter:
            keep.append(dv)
    ws.data_validations.dataValidation = keep


def apply_validation(ws, col_letter, first_data_row, named_range, growth=GROWTH_ROWS):
    """Aplica Data Validation tipo "list" sobre la columna indicada, desde
    la primera fila de datos hasta first_data_row + growth. Limpia primero
    cualquier validación previa en esa columna para mantener idempotencia.
    """
    remove_existing_validations_for_column(ws, col_letter)

    last_row = first_data_row + growth
    sqref = f"{col_letter}{first_data_row}:{col_letter}{last_row}"

    dv = DataValidation(
        type="list",
        formula1=f"={named_range}",
        allow_blank=True,
        showDropDown=False,  # mostrar el botón del dropdown
        showErrorMessage=True,
        errorTitle="ID no válido",
        error=("El valor debe coincidir con un ID existente. "
               "Usa el menú desplegable para elegir uno."),
    )
    dv.add(sqref)
    ws.add_data_validation(dv)


def main():
    print(f"[apply_dropdowns] Abriendo {XLSX} ...")
    wb = load_workbook(XLSX)

    # 1) Definir/actualizar los named ranges dinámicos.
    print("[apply_dropdowns] Configurando named ranges...")
    for name, (sheet, col_header, header_row) in LOOKUPS.items():
        ws = wb[sheet]
        col_letter = column_letter_for_header(ws, header_row, col_header)
        first_data_row = header_row + 1
        formula = build_named_range_formula(sheet, col_letter, first_data_row)
        reset_defined_name(wb, name, formula)
        print(f"  {name:18s} -> {sheet}!{col_letter} (desde fila {first_data_row}) :: {formula}")

    # 2) Aplicar data validation a cada celda referencial.
    print("[apply_dropdowns] Aplicando data validation a columnas referenciales...")
    for sheet, col_header, named_range in TARGETS:
        ws = wb[sheet]
        # Detectar header row de la hoja destino
        header_row = None
        for r in range(1, 8):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(v is not None for v in vals) and len(vals) > 1:
                header_row = r
                break
        if header_row is None:
            raise RuntimeError(f"No se detectó fila de encabezados en {sheet}")
        col_letter = column_letter_for_header(ws, header_row, col_header)
        first_data_row = header_row + 1
        apply_validation(ws, col_letter, first_data_row, named_range)
        print(f"  {sheet:24s} {col_header:18s} -> {col_letter}{first_data_row}:* (lista={named_range})")

    # 3) Guardar.
    wb.save(XLSX)
    print(f"[apply_dropdowns] OK: {XLSX}")
    print(f"[apply_dropdowns] {len(LOOKUPS)} named ranges, {len(TARGETS)} columnas con dropdown.")


if __name__ == "__main__":
    main()
