#!/usr/bin/env python3
"""Migra los IDs de 07_Investigadores del formato anónimo INV-NN al formato
con iniciales INV-XYZ derivado del nombre.

Por qué: los IDs de la forma INV-12 no le dicen nada a un editor humano que
está armando una relación en, p.ej., 13_Investigador_Modo. Con iniciales
(INV-HSG para "Herbert Spencer Gonzalez") el dropdown se vuelve
auto-explicativo y el editor reconoce a quién está vinculando.

Reglas de generación de iniciales:
  - Quitar tildes (transliteración a ASCII).
  - Tomar la primera letra de cada palabra significativa del nombre completo.
  - Saltar conectores: de, del, la, los, las, y, da, do, el, der, van, von.
    "Di" SI cuenta (tradición italiana, queda como inicial).
  - En caso de colisión (dos personas con mismas iniciales), agregar
    sufijo numerico al nuevo: HSG, HSG2, HSG3, ...

El script solo migra IDs que matcheen el patron antiguo `INV-` + digitos.
Si una fila ya tiene un ID nuevo (por ejemplo manualmente), se respeta.
Esto vuelve el script seguro de re-correr: una segunda corrida no cambia
nada.

Hojas afectadas (donde aparece el ID de investigador como dato):
  - 07_Investigadores      columna `id`        (origen)
  - 08_Temas               columna `investigador_id`
  - 12_Investigador_Lab    columna `investigador_id`
  - 13_Investigador_Modo   columna `investigador_id`

Uso:
  python3 tools/rename_investigador_ids.py            # aplica cambios
  python3 tools/rename_investigador_ids.py --dry-run  # solo muestra mapping
"""

import re
import sys
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "mad-map-data-v2.xlsx"

# Palabras que NO deben aportar inicial al ID. Conectores castellanos típicos.
SKIP_WORDS = {
    "de", "del", "la", "los", "las", "y",
    "da", "do", "el", "der", "van", "von",
}

# Patron del ID antiguo: "INV-" seguido SOLO por digitos (uno o mas).
OLD_ID_RE = re.compile(r"^INV-\d+$")

# Hojas y columnas a migrar (header_row se autodetecta).
TARGET_COLUMNS = [
    ("07_Investigadores",     "id"),
    ("08_Temas",              "investigador_id"),
    ("12_Investigador_Lab",   "investigador_id"),
    ("13_Investigador_Modo",  "investigador_id"),
]


def initials_from_name(nombre):
    """Construye el codigo de iniciales (sin prefijo INV-) a partir del nombre.
    Llamado una vez por investigador desde build_mapping().
    """
    # 1) Normalizar y eliminar marcas diacriticas (tildes).
    n = unicodedata.normalize("NFD", nombre)
    n = "".join(c for c in n if unicodedata.category(c) != "Mn")
    # 2) Separar por whitespace y filtrar conectores.
    palabras = [w for w in n.split() if w.lower() not in SKIP_WORDS]
    # 3) Tomar la primera letra de cada palabra significativa.
    return "".join(w[0].upper() for w in palabras if w)


def detect_header_row(ws, max_scan=8):
    """Detecta la fila de encabezados en una hoja. Misma heuristica que el
    resto de scripts del proyecto: primera fila donde todas las columnas
    tienen valor no-vacio.
    """
    for r in range(1, max_scan + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is not None for v in vals) and len(vals) > 1:
            return r
    raise RuntimeError(f"No se detecto fila de encabezados en hoja {ws.title!r}")


def column_letter_for_header(ws, header_row, header_name):
    """Devuelve la letra de columna cuya celda en header_row matchea
    header_name. Lanza ValueError si no la encuentra.
    """
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row, c).value == header_name:
            return ws.cell(header_row, c).column_letter
    raise ValueError(
        f"No se encontro la columna {header_name!r} en hoja {ws.title!r}"
    )


def build_mapping(wb):
    """Lee 07_Investigadores y construye {old_id: new_id} solo para los IDs
    con formato antiguo. Resuelve colisiones agregando sufijo numerico.
    Llamado una vez al inicio de main().
    """
    ws = wb["07_Investigadores"]
    header_row = detect_header_row(ws)
    id_col = column_letter_for_header(ws, header_row, "id")
    nombre_col = column_letter_for_header(ws, header_row, "nombre")

    # Recolectar IDs ya en formato nuevo (o cualquier no-antiguo) para no
    # generar colisiones contra ellos.
    used = set()
    for r in range(header_row + 1, ws.max_row + 1):
        old_id = ws[f"{id_col}{r}"].value
        if old_id and not OLD_ID_RE.match(str(old_id)):
            used.add(old_id)

    mapping = {}
    for r in range(header_row + 1, ws.max_row + 1):
        old_id = ws[f"{id_col}{r}"].value
        nombre = ws[f"{nombre_col}{r}"].value
        if not old_id or not nombre:
            continue
        if not OLD_ID_RE.match(str(old_id)):
            continue  # ya migrado, respetar.
        base = f"INV-{initials_from_name(nombre)}"
        candidate = base
        suffix = 2
        while candidate in used:
            candidate = f"{base}{suffix}"
            suffix += 1
        used.add(candidate)
        mapping[old_id] = candidate
    return mapping


def apply_mapping(wb, mapping):
    """Reemplaza los valores de las celdas que matcheen una clave del mapping
    en todas las TARGET_COLUMNS. Llamado una vez en main(), despues de
    build_mapping().

    Devuelve un dict {sheet: count} con cuantas celdas se actualizaron.
    """
    counts = {}
    for sheet, header_name in TARGET_COLUMNS:
        ws = wb[sheet]
        header_row = detect_header_row(ws)
        col_letter = column_letter_for_header(ws, header_row, header_name)
        col_idx = ws[f"{col_letter}{header_row}"].column
        n = 0
        for r in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(r, col_idx)
            if cell.value in mapping:
                cell.value = mapping[cell.value]
                n += 1
        counts[sheet] = n
    return counts


def main(dry_run=False):
    print(f"[rename_inv_ids] Abriendo {XLSX} ...")
    wb = load_workbook(XLSX)

    mapping = build_mapping(wb)
    if not mapping:
        print("[rename_inv_ids] No hay IDs en formato antiguo. Nada que hacer.")
        return

    print("[rename_inv_ids] Mapping de IDs:")
    # Usar el orden de aparicion en 07_Investigadores
    ws = wb["07_Investigadores"]
    header_row = detect_header_row(ws)
    id_col = column_letter_for_header(ws, header_row, "id")
    nombre_col = column_letter_for_header(ws, header_row, "nombre")
    for r in range(header_row + 1, ws.max_row + 1):
        old_id = ws[f"{id_col}{r}"].value
        nombre = ws[f"{nombre_col}{r}"].value
        if old_id in mapping:
            print(f"  {old_id} -> {mapping[old_id]:10s} ({nombre})")

    if dry_run:
        print("\n[rename_inv_ids] dry-run: no se aplicaron cambios.")
        return

    counts = apply_mapping(wb, mapping)
    print("[rename_inv_ids] Celdas actualizadas por hoja:")
    for sheet, n in counts.items():
        print(f"  {sheet:32s} {n:4d}")

    wb.save(XLSX)
    print(f"[rename_inv_ids] OK: {XLSX}")


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    main(dry_run=dry)
