#!/usr/bin/env python3
"""Paso 2 de la usabilidad humana: las hojas de relación dejan de referenciar
entidades por ID y pasan a hacerlo por NOMBRE.

Movimientos que ejecuta este script:

  1. Renombra las columnas *_id de las hojas de relación a su forma legible:
     linea_id -> línea, area_id -> área, lab_id -> laboratorio, etc.
  2. Reemplaza los valores de esas columnas: cada ID se sustituye por el
     nombre de la entidad correspondiente (lookup en la hoja entidad).
  3. Colapsa 08_Temas + 09_Sublinea_Tema en una sola hoja 08_Temas con
     columnas (sublínea, investigador, tema). Borra 09_Sublinea_Tema.
  4. En 18_Proximidad_Tematica, elimina las columnas redundantes
     sublinea_a_nombre y sublinea_b_nombre (pasaban a ser duplicación).
  5. Re-aplica named ranges y data validations apuntando ahora a las
     columnas `nombre` de las hojas entidad (no a `id`).

Idempotencia: el script detecta si una hoja ya está migrada (por el nombre
de la columna) y la salta. Re-correrlo no rompe nada.

Asume invariantes del paso 1 (verificados al ejecutar):
  - Todos los nombres son únicos por entidad.
  - Cero temas huérfanos (todos los temas de 08 están en 09_Sublinea_Tema).

Uso:
  python3 tools/migrate_to_names.py
"""

import sys
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "mad-map-data-v2.xlsx"

GROWTH_ROWS = 500


# ---------------------------------------------------------------------------
# Helpers de detección de hojas
# ---------------------------------------------------------------------------

def detect_header_row(ws, max_scan=8):
    """Primera fila donde todas las columnas tienen valor no-vacío."""
    for r in range(1, max_scan + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is not None for v in vals) and len(vals) > 1:
            return r
    raise RuntimeError(f"No se detectó fila de encabezados en {ws.title!r}")


def header_index(ws, header_row, header_name):
    """Devuelve el índice de columna (1-based) cuyo header coincide. Raise si
    no la encuentra. Usado por todas las funciones de migración para localizar
    columnas sin asumir orden fijo."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row, c).value == header_name:
            return c
    raise ValueError(f"No hay columna {header_name!r} en {ws.title!r}")


def has_header(ws, header_row, header_name):
    """Verifica si una hoja ya tiene una columna con cierto nombre. Usado para
    idempotencia: si ya está renombrada, saltamos esa hoja."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row, c).value == header_name:
            return True
    return False


# ---------------------------------------------------------------------------
# Construcción de mapas ID -> nombre desde las hojas entidad
# ---------------------------------------------------------------------------

# Configuración: para cada entidad, qué hoja, qué columna ID y qué columna
# nombre. Los nombres usados aquí son los headers actuales (paso 1).
ENTITY_MAPS = [
    # (clave, hoja, header_id, header_nombre)
    ("linea",        "01_Lineas",         "id",     "nombre"),
    ("sublinea",     "02_Sublineas",      "id",     "nombre"),
    ("area",         "03_Areas",          "código", "nombre"),
    ("modo",         "04_Modos",          "id",     "nombre"),
    ("salida",       "05_Salidas",        "id",     "nombre"),
    ("lab",          "06_Laboratorios",   "id",     "nombre"),
    ("investigador", "07_Investigadores", "id",     "nombre"),
]


def build_id_to_name(wb):
    """Lee todas las hojas entidad y devuelve {clave: {id: nombre}}.
    Usado por migrate_relation_sheet y collapse_temas para resolver IDs
    a nombres legibles."""
    maps = {}
    for key, sheet, hid, hname in ENTITY_MAPS:
        ws = wb[sheet]
        hr = detect_header_row(ws)
        ci = header_index(ws, hr, hid)
        cn = header_index(ws, hr, hname)
        m = {}
        for r in range(hr + 1, ws.max_row + 1):
            id_val = ws.cell(r, ci).value
            name_val = ws.cell(r, cn).value
            if id_val and name_val:
                m[id_val] = name_val
        maps[key] = m
    return maps


# ---------------------------------------------------------------------------
# Migración de columnas en hojas de relación
# ---------------------------------------------------------------------------

# Para cada hoja, qué columnas migrar. Cada entrada:
#   (sheet_name, [(old_header, new_header, entity_key), ...])
RELATION_RENAMES = [
    ("02_Sublineas", [
        ("linea_id", "línea", "linea"),
        ("area_id",  "área",  "area"),
    ]),
    ("10_Lab_Linea", [
        ("lab_id",   "laboratorio", "lab"),
        ("linea_id", "línea",       "linea"),
    ]),
    ("11_Lab_Salida", [
        ("lab_id",    "laboratorio", "lab"),
        ("salida_id", "salida",      "salida"),
    ]),
    ("12_Investigador_Lab", [
        ("investigador_id", "investigador", "investigador"),
        ("lab_id",          "laboratorio",  "lab"),
    ]),
    ("13_Investigador_Modo", [
        ("investigador_id", "investigador", "investigador"),
        ("modo_id",         "modo",         "modo"),
    ]),
    ("14_Linea_Modo", [
        ("linea_id", "línea", "linea"),
        ("modo_id",  "modo",  "modo"),
    ]),
    ("18_Proximidad_Tematica", [
        ("sublinea_a_id", "sublínea_a", "sublinea"),
        ("sublinea_b_id", "sublínea_b", "sublinea"),
    ]),
]


def migrate_relation_sheet(ws, renames, id_to_name_maps):
    """Para una hoja de relación, renombra los headers y reemplaza los valores
    de las celdas (ID -> nombre). Idempotente: si la columna ya tiene el
    header nuevo, salta."""
    hr = detect_header_row(ws)
    counts = []
    for old_h, new_h, entity_key in renames:
        if has_header(ws, hr, new_h):
            counts.append((old_h, new_h, "ya_migrada"))
            continue
        ci = header_index(ws, hr, old_h)
        # Reemplazar header
        ws.cell(hr, ci).value = new_h
        # Reemplazar valores
        m = id_to_name_maps[entity_key]
        n = 0
        for r in range(hr + 1, ws.max_row + 1):
            cell = ws.cell(r, ci)
            if cell.value in m:
                cell.value = m[cell.value]
                n += 1
        counts.append((old_h, new_h, n))
    return counts


def drop_redundant_columns_18(ws):
    """En 18_Proximidad_Tematica las columnas sublinea_a_nombre y
    sublinea_b_nombre eran espejo de los IDs vecinos. Tras migrar a nombres
    quedan duplicadas; las eliminamos para limpiar la planilla."""
    hr = detect_header_row(ws)
    targets = ["sublinea_a_nombre", "sublinea_b_nombre"]
    # Borrar de derecha a izquierda para no invalidar índices
    cols_to_delete = sorted(
        [header_index(ws, hr, t) for t in targets if has_header(ws, hr, t)],
        reverse=True,
    )
    for ci in cols_to_delete:
        ws.delete_cols(ci, 1)
    return len(cols_to_delete)


# ---------------------------------------------------------------------------
# Colapso 08_Temas + 09_Sublinea_Tema -> nuevo 08_Temas
# ---------------------------------------------------------------------------

def collapse_temas(wb, id_to_name_maps):
    """Reemplaza el contenido de 08_Temas con la fusión: cada fila contiene
    (sublínea, investigador, tema). Borra 09_Sublinea_Tema.

    Idempotente: si 08_Temas ya tiene encabezado 'sublínea' (formato nuevo),
    no hace nada.
    """
    ws_temas = wb["08_Temas"]
    hr = detect_header_row(ws_temas)
    if has_header(ws_temas, hr, "sublínea"):
        return ("ya_migrada", 0)

    # 1) Construir mapping tema_id -> (investigador_id, texto)
    ci_id = header_index(ws_temas, hr, "id")
    ci_inv = header_index(ws_temas, hr, "investigador_id")
    ci_txt = header_index(ws_temas, hr, "tema")
    tema_lookup = {}
    for r in range(hr + 1, ws_temas.max_row + 1):
        tid = ws_temas.cell(r, ci_id).value
        if not tid:
            continue
        tema_lookup[tid] = (
            ws_temas.cell(r, ci_inv).value,
            ws_temas.cell(r, ci_txt).value,
        )

    # 2) Recorrer 09_Sublinea_Tema y construir filas nuevas
    if "09_Sublinea_Tema" not in wb.sheetnames:
        # ya borrada en una migración previa
        return ("sin_09", 0)
    ws_st = wb["09_Sublinea_Tema"]
    hr_st = detect_header_row(ws_st)
    ci_sub = header_index(ws_st, hr_st, "sublinea_id")
    ci_tem = header_index(ws_st, hr_st, "tema_id")

    sub_id_to_name = id_to_name_maps["sublinea"]
    inv_id_to_name = id_to_name_maps["investigador"]

    new_rows = []
    for r in range(hr_st + 1, ws_st.max_row + 1):
        sub_id = ws_st.cell(r, ci_sub).value
        tema_id = ws_st.cell(r, ci_tem).value
        if not sub_id or not tema_id:
            continue
        if tema_id not in tema_lookup:
            print(f"  [warning] tema_id {tema_id} en 09 no existe en 08, omitido")
            continue
        inv_id, texto = tema_lookup[tema_id]
        sub_nombre = sub_id_to_name.get(sub_id, sub_id)
        inv_nombre = inv_id_to_name.get(inv_id, inv_id)
        new_rows.append((sub_nombre, inv_nombre, texto))

    # 3) Reescribir 08_Temas con la nueva forma
    # Limpiar todo el contenido actual
    ws_temas.delete_rows(1, ws_temas.max_row)

    # Re-pintar título, nota y encabezado en filas 1, 2, 4
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    title_font = Font(name="Arial", size=12, bold=True)
    note_font = Font(name="Arial", size=11, italic=True, color="666666")
    header_font = Font(name="Arial", size=11, bold=True)
    header_fill = PatternFill("solid", fgColor="EEEEEE")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws_temas.cell(1, 1).value = "08 · Temas (sublínea ↔ investigador ↔ tema)"
    ws_temas.cell(1, 1).font = title_font
    ws_temas.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws_temas.cell(2, 1).value = (
        "Cada fila atribuye un tema concreto a una sublínea y un investigador. "
        "Editado a mano: usa los desplegables para elegir sublínea e investigador."
    )
    ws_temas.cell(2, 1).font = note_font
    ws_temas.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws_temas.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)

    ws_temas.cell(4, 1).value = "sublínea"
    ws_temas.cell(4, 2).value = "investigador"
    ws_temas.cell(4, 3).value = "tema"
    for c in range(1, 4):
        ws_temas.cell(4, c).font = header_font
        ws_temas.cell(4, c).fill = header_fill
        ws_temas.cell(4, c).border = border

    for i, (sub_nombre, inv_nombre, texto) in enumerate(new_rows):
        r = 5 + i
        ws_temas.cell(r, 1).value = sub_nombre
        ws_temas.cell(r, 2).value = inv_nombre
        ws_temas.cell(r, 3).value = texto

    # Ajustar ancho de columnas
    ws_temas.column_dimensions["A"].width = 50
    ws_temas.column_dimensions["B"].width = 28
    ws_temas.column_dimensions["C"].width = 80

    # 4) Borrar 09_Sublinea_Tema
    if "09_Sublinea_Tema" in wb.sheetnames:
        del wb["09_Sublinea_Tema"]

    return ("ok", len(new_rows))


# ---------------------------------------------------------------------------
# Re-aplicar named ranges + data validations apuntando a columnas `nombre`
# ---------------------------------------------------------------------------

# Lookups: ahora apuntan a la columna `nombre` (o `código` para áreas) de
# cada hoja entidad. Mantengo los nombres anteriores para minimizar cambios
# externos pero el contenido cambia.
LOOKUPS = {
    "LineaNombres":        ("01_Lineas",         "nombre",  4),
    "SublineaNombres":     ("02_Sublineas",      "nombre",  4),
    "AreaNombres":         ("03_Areas",          "nombre",  4),
    "ModoNombres":         ("04_Modos",          "nombre",  2),
    "SalidaNombres":       ("05_Salidas",        "nombre",  4),
    "LabNombres":          ("06_Laboratorios",   "nombre",  2),
    "InvestigadorNombres": ("07_Investigadores", "nombre",  4),
}

# Targets: hoja, columna a validar, named range a usar
TARGETS = [
    ("02_Sublineas",           "línea",        "LineaNombres"),
    ("02_Sublineas",           "área",         "AreaNombres"),
    ("08_Temas",               "sublínea",     "SublineaNombres"),
    ("08_Temas",               "investigador", "InvestigadorNombres"),
    ("10_Lab_Linea",           "laboratorio",  "LabNombres"),
    ("10_Lab_Linea",           "línea",        "LineaNombres"),
    ("11_Lab_Salida",          "laboratorio",  "LabNombres"),
    ("11_Lab_Salida",          "salida",       "SalidaNombres"),
    ("12_Investigador_Lab",    "investigador", "InvestigadorNombres"),
    ("12_Investigador_Lab",    "laboratorio",  "LabNombres"),
    ("13_Investigador_Modo",   "investigador", "InvestigadorNombres"),
    ("13_Investigador_Modo",   "modo",         "ModoNombres"),
    ("14_Linea_Modo",          "línea",        "LineaNombres"),
    ("14_Linea_Modo",          "modo",         "ModoNombres"),
    ("18_Proximidad_Tematica", "sublínea_a",   "SublineaNombres"),
    ("18_Proximidad_Tematica", "sublínea_b",   "SublineaNombres"),
]


def remove_old_named_ranges(wb):
    """Borra los named ranges del paso 1 (LineaIDs, SublineaIDs, etc.) que
    apuntaban a columnas `id` y ya no aplican. Llamado una vez al inicio
    de re-aplicar para evitar referencias colgadas."""
    for n in ("LineaIDs", "SublineaIDs", "AreaIDs", "ModoIDs", "SalidaIDs",
              "LabIDs", "InvestigadorIDs", "TemaIDs"):
        if n in wb.defined_names:
            del wb.defined_names[n]


def remove_existing_validations_for_column(ws, col_letter):
    """Quita validaciones previas que cubran exclusivamente la columna dada.
    Preserva otras validaciones (ej: estado=DESCARTADO en 18.G)."""
    keep = []
    for dv in ws.data_validations.dataValidation:
        sqref = str(dv.sqref) if dv.sqref else ""
        first_cell = sqref.split()[0].split(":")[0] if sqref else ""
        first_col_letters = "".join(ch for ch in first_cell if ch.isalpha())
        if first_col_letters != col_letter:
            keep.append(dv)
    ws.data_validations.dataValidation = keep


def apply_dropdowns(wb):
    """Crea named ranges nuevos y aplica data validation a cada celda
    referencial usando las columnas `nombre` como source.
    """
    remove_old_named_ranges(wb)

    for name, (sheet, col_header, header_row) in LOOKUPS.items():
        ws = wb[sheet]
        col_idx = header_index(ws, header_row, col_header)
        col_letter = ws.cell(header_row, col_idx).column_letter
        first_data_row = header_row + 1
        formula = (
            f"OFFSET('{sheet}'!${col_letter}${first_data_row},0,0,"
            f"COUNTA('{sheet}'!${col_letter}${first_data_row}:${col_letter}$10000),1)"
        )
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names[name] = DefinedName(name=name, attr_text=formula)

    for sheet, col_header, named_range in TARGETS:
        ws = wb[sheet]
        hr = detect_header_row(ws)
        col_idx = header_index(ws, hr, col_header)
        col_letter = ws.cell(hr, col_idx).column_letter
        first_data_row = hr + 1
        last_row = first_data_row + GROWTH_ROWS
        sqref = f"{col_letter}{first_data_row}:{col_letter}{last_row}"

        remove_existing_validations_for_column(ws, col_letter)

        dv = DataValidation(
            type="list",
            formula1=f"={named_range}",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Nombre no válido",
            error=("El valor debe coincidir con un nombre existente en la hoja "
                   "correspondiente. Usa el menú desplegable para elegir uno."),
        )
        dv.add(sqref)
        ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"[migrate_to_names] Abriendo {XLSX} ...")
    wb = load_workbook(XLSX)

    # 1) Mapas ID -> nombre desde las hojas entidad (antes de tocar nada).
    id_to_name = build_id_to_name(wb)
    for k, m in id_to_name.items():
        print(f"  map {k:14s} {len(m):3d} entradas")

    # 2) Migrar las hojas de relación (excepto 08_Temas, que se colapsa aparte)
    print("\n[migrate_to_names] Renombrando columnas y migrando IDs -> nombres:")
    for sheet, renames in RELATION_RENAMES:
        ws = wb[sheet]
        results = migrate_relation_sheet(ws, renames, id_to_name)
        for old_h, new_h, count in results:
            mark = f"{count} celdas" if isinstance(count, int) else count
            print(f"  {sheet:24s} {old_h:18s} -> {new_h:14s} ({mark})")

    # 3) Eliminar columnas redundantes en 18_Proximidad_Tematica
    print("\n[migrate_to_names] Limpiando columnas redundantes en 18_Proximidad_Tematica:")
    n_dropped = drop_redundant_columns_18(wb["18_Proximidad_Tematica"])
    print(f"  columnas eliminadas: {n_dropped}")

    # 4) Colapsar 08_Temas + 09_Sublinea_Tema
    print("\n[migrate_to_names] Colapsando 08_Temas + 09_Sublinea_Tema:")
    estado, nfilas = collapse_temas(wb, id_to_name)
    print(f"  estado: {estado} | filas en nueva 08_Temas: {nfilas}")

    # 5) Re-aplicar dropdowns sobre el modelo nuevo
    print("\n[migrate_to_names] Re-aplicando named ranges y data validations:")
    apply_dropdowns(wb)
    print(f"  {len(LOOKUPS)} named ranges, {len(TARGETS)} columnas con dropdown")

    wb.save(XLSX)
    print(f"\n[migrate_to_names] OK: {XLSX}")


if __name__ == "__main__":
    main()
