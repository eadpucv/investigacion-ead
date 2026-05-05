#!/usr/bin/env python3
"""Repara errores y drift detectados en mad-map-data-v2.xlsx tras el paso 2.

Hallazgos que arregla:

  1. Hoja 16_Resumen tiene fórmulas que apuntan a la hoja 09_Sublinea_Tema
     (colapsada en 08_Temas durante migrate_to_names.py). Esto genera el
     aviso "Es posible que el archivo se haya movido..." al abrir el .xlsx.
     Acción: actualizar las fórmulas para que cuenten desde 08_Temas en
     su nueva forma (sublínea, investigador, tema).

  2. Hoja 00_Lectura tiene una mención textual a "09_Sublinea_Tema" en una
     instrucción descriptiva. Acción: actualizar el texto para describir
     el modelo nuevo.

  3. Hoja 18_Proximidad_Tematica reporta max_column=7 pese a que el header
     efectivo termina en la columna 5. Son artefactos de un delete_cols
     anterior que dejó datos vacíos en columnas espurias. Acción: rebuild
     de la hoja recortando las columnas fantasma.

  4. Auditoría reporta que los 86 pares de proximidad están todos en
     dirección única (no simétricos en pareja). Acción: NO hace falta
     duplicar — la decisión de modelo es tratar la proximidad como
     representación canónica (un par por relación). Documento en el roadmap
     que el invariante ProximidadSimetrica del spec original se reemplaza
     por ProximidadCanonica.

Idempotente: detecta si las correcciones ya fueron aplicadas y salta.
"""

import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "mad-map-data-v2.xlsx"


def remove_external_link(xlsx_path):
    """El xlsx tiene un external link colgado a una hoja inexistente
    ('09_Sublinea_Tema'), generado automáticamente por Excel cuando se
    eliminó la hoja referenciada por una fórmula. Esto produce el aviso
    'Es posible que el archivo se haya movido' al abrir el archivo.

    Limpiamos el external link a nivel de empaquetado XML del .xlsx (que
    es un .zip), porque openpyxl no expone API para esto. Operamos sobre
    una copia en memoria y reescribimos el zip.

    Movimientos:
      - Quitar xl/externalLinks/externalLink1.xml y su .rels.
      - Quitar el <externalReference> de xl/workbook.xml.
      - Quitar la <Relationship type="externalLink"> de workbook.xml.rels.
      - Quitar el <Override> correspondiente de [Content_Types].xml.
      - En xl/worksheets/sheet16.xml, reescribir fórmulas que decían
        '[1]09_Sublinea_Tema'!... por '08_Temas'!... con un cómputo
        equivalente (cuenta atribuciones tema↔sublínea, que es lo mismo
        que producía la hoja vieja).

    Idempotente: si los archivos ya no están, no hace nada.
    """
    import shutil
    import zipfile

    tmp_path = xlsx_path.with_suffix(".xlsx.tmp")
    shutil.copy(xlsx_path, tmp_path)

    with zipfile.ZipFile(tmp_path) as src:
        members = {n: src.read(n) for n in src.namelist()}

    # 1) Detectar si hay external link al que apuntar
    if "xl/externalLinks/externalLink1.xml" not in members:
        tmp_path.unlink()
        return ("ya_limpio", 0)

    # 2) Eliminar archivos del external link
    removed_paths = []
    for path in (
        "xl/externalLinks/externalLink1.xml",
        "xl/externalLinks/_rels/externalLink1.xml.rels",
    ):
        if path in members:
            del members[path]
            removed_paths.append(path)

    # 3) Quitar <externalReferences>...</externalReferences> de workbook.xml
    wb_xml = members.get("xl/workbook.xml", b"").decode("utf-8")
    wb_xml = re.sub(
        r"<externalReferences>.*?</externalReferences>", "", wb_xml, flags=re.S
    )
    members["xl/workbook.xml"] = wb_xml.encode("utf-8")

    # 4) Quitar la Relationship del externalLink en workbook.xml.rels
    rels_xml = members.get("xl/_rels/workbook.xml.rels", b"").decode("utf-8")
    rels_xml = re.sub(
        r'<Relationship\s+Type="[^"]*externalLink[^"]*"[^/]*/>', "", rels_xml
    )
    members["xl/_rels/workbook.xml.rels"] = rels_xml.encode("utf-8")

    # 5) Quitar el Override de [Content_Types].xml
    ct_xml = members.get("[Content_Types].xml", b"").decode("utf-8")
    ct_xml = re.sub(
        r'<Override\s+PartName="/xl/externalLinks/externalLink1\.xml"[^/]*/>',
        "",
        ct_xml,
    )
    members["[Content_Types].xml"] = ct_xml.encode("utf-8")

    # 6) Reescribir fórmulas en sheet16.xml: cambiar referencias a la hoja
    #    inexistente por referencias directas a 08_Temas con cómputo
    #    equivalente. Las dos fórmulas relevantes:
    #      a) <f>COUNTA('[1]09_Sublinea_Tema'!A:A)-1</f>
    #         que contaba pares sublínea-tema. Ahora cada fila de 08_Temas
    #         es exactamente una atribución, así que:
    #         <f>COUNTA('08_Temas'!A:A)-3</f>
    #         (resta 3: título, nota y fila de header — que ahora ocupa
    #         hasta la fila 4 inclusive en 08_Temas).
    #      b) <f>SUMPRODUCT(1/COUNTIF('[1]09_Sublinea_Tema'!B2:B500,
    #            '[1]09_Sublinea_Tema'!B2:B500))</f>
    #         contaba sublíneas únicas con tema. Equivalente:
    #         <f>SUMPRODUCT((A5:A500<>"")/COUNTIF(A5:A500,A5:A500&""))</f>
    #         sobre 08_Temas.A.
    sh16_xml = members.get("xl/worksheets/sheet16.xml", b"").decode("utf-8")
    sh16_xml = sh16_xml.replace(
        "COUNTA('[1]09_Sublinea_Tema'!A:A)-1",
        "COUNTA('08_Temas'!A:A)-3",
    )
    # Reescribir el SUMPRODUCT que contaba sublíneas únicas. Importante:
    # esto va dentro de XML (<f>...</f>), por lo que cualquier caracter
    # reservado debe escaparse: < > & " se vuelven &lt; &gt; &amp; &quot;.
    # La fórmula efectiva en Excel es:
    #   SUMPRODUCT(('08_Temas'!A5:A500<>"")/COUNTIF('08_Temas'!A5:A500,
    #              '08_Temas'!A5:A500&""))
    new_sumproduct = (
        "SUMPRODUCT(('08_Temas'!A5:A500&lt;&gt;&quot;&quot;)"
        "/COUNTIF('08_Temas'!A5:A500,'08_Temas'!A5:A500&amp;&quot;&quot;))"
    )
    sh16_xml = re.sub(
        r"SUMPRODUCT\(1/COUNTIF\('\[1\]09_Sublinea_Tema'!B2:B500,'\[1\]09_Sublinea_Tema'!B2:B500\)\)",
        new_sumproduct,
        sh16_xml,
    )
    members["xl/worksheets/sheet16.xml"] = sh16_xml.encode("utf-8")

    # 7) Reescribir el zip con los cambios. Abrimos el destino con 'wb'
    #    para truncar in-place (algunos filesystems prohíben unlink pero
    #    permiten truncate). ZipFile sobre el file-object funciona igual.
    with open(xlsx_path, "wb") as fh:
        with zipfile.ZipFile(fh, "w", zipfile.ZIP_DEFLATED) as dst:
            for name, content in members.items():
                dst.writestr(name, content)
    # Borrar el .tmp si el filesystem lo permite; si no, queda como
    # backup auto-creado por shutil.copy y es inofensivo.
    try:
        tmp_path.unlink()
    except OSError:
        pass
    return ("ok", len(removed_paths))


def fix_lectura_text(wb):
    """00_Lectura menciona la hoja vieja 09_Sublinea_Tema. Refrescamos
    el texto para describir la forma actual de 08_Temas."""
    ws = wb["00_Lectura"]
    fixed = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if not isinstance(v, str) or "09_Sublinea_Tema" not in v:
                continue
            # Reemplazar la instrucción de edición.
            new_v = v.replace(
                "Editar en la hoja 09_Sublinea_Tema",
                "Editar en la hoja 08_Temas (cada fila atribuye un tema "
                "a una sublínea y un investigador)",
            )
            if new_v != v:
                cell.value = new_v
                fixed += 1
    return fixed


def trim_phantom_columns_18(wb):
    """18_Proximidad_Tematica tiene columnas fantasma al final. Las
    detectamos buscando la última columna donde hay header válido (fila 4),
    y eliminamos las columnas posteriores si quedan vacías por completo."""
    ws = wb["18_Proximidad_Tematica"]
    header_row = 4
    # Encontrar última columna con header no vacío
    last_real = 0
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row, c).value is not None:
            last_real = c
    # Verificar que las columnas posteriores están realmente vacías
    cols_to_delete = []
    for c in range(last_real + 1, ws.max_column + 1):
        all_empty = all(
            ws.cell(r, c).value is None
            for r in range(1, ws.max_row + 1)
        )
        if all_empty:
            cols_to_delete.append(c)
    # Borrar de derecha a izquierda
    for c in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(c, 1)
    return len(cols_to_delete)


def report_proximity_canonicalization(wb):
    """Reporta cuántos pares hay en proximidad y verifica si la
    representación es canónica (sin duplicados a↔b/b↔a). No modifica."""
    ws = wb["18_Proximidad_Tematica"]
    header_row = 4
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    while headers and headers[-1] is None:
        headers.pop()
    col_a = headers.index("sublínea_a") + 1
    col_b = headers.index("sublínea_b") + 1
    col_estado = headers.index("estado") + 1 if "estado" in headers else None

    seen = set()
    duplicates = 0
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(r, col_a).value
        b = ws.cell(r, col_b).value
        if not (a and b):
            continue
        if col_estado and ws.cell(r, col_estado).value == "DESCARTADO":
            continue
        key = tuple(sorted([a, b]))
        if key in seen:
            duplicates += 1
        seen.add(key)
    return len(seen), duplicates


def main():
    print(f"[clean_xlsx] Abriendo {XLSX} ...")

    # Pasada 1: usando openpyxl (modificaciones a nivel de celda)
    wb = load_workbook(XLSX)

    print("\n[1] Refrescando texto de 00_Lectura:")
    n = fix_lectura_text(wb)
    print(f"    celdas de texto actualizadas: {n}")

    print("\n[2] Recortando columnas fantasma de 18_Proximidad_Tematica:")
    n = trim_phantom_columns_18(wb)
    print(f"    columnas eliminadas: {n}")

    print("\n[3] Estado actual de la matriz de proximidad:")
    canon_count, dup_count = report_proximity_canonicalization(wb)
    print(f"    pares únicos canónicos: {canon_count}")
    print(f"    duplicados a↔b/b↔a:     {dup_count}")
    if dup_count == 0:
        print("    -> la representación ya es canónica (un par por relación).")

    wb.save(XLSX)

    # Pasada 2: a nivel de empaquetado XML del .xlsx (zipfile), porque
    # openpyxl no expone API para eliminar external links.
    print("\n[4] Eliminando external link colgado a 09_Sublinea_Tema:")
    estado, n_paths = remove_external_link(XLSX)
    if estado == "ya_limpio":
        print("    no hay external links, nada que hacer.")
    else:
        print(f"    archivos XML eliminados del paquete: {n_paths}")
        print(f"    fórmulas en sheet16.xml reescritas a apuntar a 08_Temas")

    print(f"\n[clean_xlsx] OK: {XLSX}")


if __name__ == "__main__":
    main()
