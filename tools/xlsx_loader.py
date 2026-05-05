"""Equivalente Python del xlsx-loader.js: lee mad-map-data-v2.xlsx y devuelve
el modelo del grafo (entidades, relaciones, aristas) que consume cualquier
herramienta interna que necesite los datos del mapa.

Hoy lo usa tools/build_doc.py para regenerar el documento institucional
sin pasar por un JSON intermedio.

La estructura devuelta es la misma que producía la versión vieja basada en
mad-map-data.json, salvo que ya no incluye `position` (las coordenadas del
grafo se calculan al vuelo en el navegador con D3 force-directed).
"""

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Helpers de lectura
# ---------------------------------------------------------------------------

def _detect_header_row(ws, max_scan=8):
    """Heurística para detectar la fila de encabezados: la primera fila
    donde, ignorando celdas None al final (resto de delete_cols antiguos),
    quedan al menos 2 celdas no vacías sin huecos entre ellas.

    Llamado una vez por hoja, en _read_sheet.
    """
    for r in range(1, max_scan + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        # Recortar Nones al final (artefactos de columnas eliminadas)
        while vals and vals[-1] is None:
            vals.pop()
        if len(vals) >= 2 and all(v is not None for v in vals):
            return r
    raise RuntimeError(f"No se detectó fila de encabezados en {ws.title!r}")


def _read_sheet(wb, sheet_name):
    """Lee una hoja del workbook y devuelve [{col: valor}, ...].
    Salta la fila de encabezados detectada automáticamente.
    """
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    hr = _detect_header_row(ws)
    headers = [ws.cell(hr, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(hr + 1, ws.max_row + 1):
        row = {}
        empty = True
        for i, h in enumerate(headers, start=1):
            if h is None:
                continue
            v = ws.cell(r, i).value
            row[h] = v
            if v is not None and str(v).strip() != "":
                empty = False
        if not empty:
            rows.append(row)
    return rows


def _build_name_to_id(entities, label):
    """Construye {nombre: id} para una lista de entidades. Reporta duplicados
    en stderr (por estilo, sin abortar). Llamado por load() para cada tipo."""
    import sys
    m = {}
    seen = set()
    for e in entities:
        if not e.get("nombre"):
            continue
        key = str(e["nombre"]).strip()
        if key in seen:
            print(f"[xlsx_loader] WARNING nombre duplicado en {label}: {key!r}",
                  file=sys.stderr)
            continue
        seen.add(key)
        m[key] = e["id"]
    return m


def _resolve(value, name_to_id, context):
    """Resuelve un nombre a su ID interno usando el mapa. Si no resuelve,
    emite warning y devuelve None. Igual que la versión JS."""
    import sys
    if value is None or value == "":
        return None
    key = str(value).strip()
    if key in name_to_id:
        return name_to_id[key]
    print(f"[xlsx_loader] WARNING referencia no resuelta en {context}: {key!r}",
          file=sys.stderr)
    return None


# ---------------------------------------------------------------------------
# Cómputo de aristas (mismas 7 categorías que xlsx-loader.js)
# ---------------------------------------------------------------------------

def _compute_edges(*, sublineas, lab_linea, linea_modo, proximidad,
                   sublinea_tema, temas):
    """Reproduce computeEdges del loader JS."""
    edges = {
        "jerarquica": [],
        "coautoria": [],
        "coinvestigacion": [],
        "sosten_lab": [],
        "afinidad_lab": [],
        "coincidencia_modo": [],
        "proximidad_semantica": [],
    }

    # (a) Jerárquica: cada Sublínea → Línea
    for s in sublineas:
        if s.get("linea"):
            edges["jerarquica"].append({"source": s["id"], "target": s["linea"]})

    # (b) Coautoría: Investigador → Sublínea, vía Tema
    tema_to_inv = {t["id"]: t["investigador"] for t in temas}
    inv_to_subs = defaultdict(set)
    for st in sublinea_tema:
        inv_id = tema_to_inv.get(st.get("tema_id"))
        if inv_id:
            inv_to_subs[inv_id].add(st["sublinea_id"])
    for inv_id, subs in inv_to_subs.items():
        for sub_id in subs:
            edges["coautoria"].append({"source": inv_id, "target": sub_id})

    # (c) Coinvestigación: pares con investigador compartido
    sub_to_invs = defaultdict(set)
    for inv_id, subs in inv_to_subs.items():
        for sub_id in subs:
            sub_to_invs[sub_id].add(inv_id)
    sub_ids = sorted(sub_to_invs.keys())
    for i, a in enumerate(sub_ids):
        for b in sub_ids[i + 1:]:
            shared = sub_to_invs[a] & sub_to_invs[b]
            if shared:
                edges["coinvestigacion"].append({
                    "source": a, "target": b, "weight": len(shared),
                })

    # (d) Sostén Lab → Línea
    for ll in lab_linea:
        edges["sosten_lab"].append({"source": ll["lab_id"], "target": ll["linea_id"]})

    # (e) Afinidad por Lab: pares de sublíneas con lab compartido
    linea_to_labs = defaultdict(set)
    for ll in lab_linea:
        linea_to_labs[ll["linea_id"]].add(ll["lab_id"])
    sub_to_labs = {}
    for s in sublineas:
        if s.get("linea") and s["linea"] in linea_to_labs:
            sub_to_labs[s["id"]] = linea_to_labs[s["linea"]]
    sub_ids_lab = sorted(sub_to_labs.keys())
    for i, a in enumerate(sub_ids_lab):
        for b in sub_ids_lab[i + 1:]:
            if sub_to_labs[a] & sub_to_labs[b]:
                edges["afinidad_lab"].append({"source": a, "target": b})

    # (f) Coincidencia de modo (sólo nivel predominante)
    linea_to_modos = defaultdict(set)
    for lm in linea_modo:
        if str(lm.get("nivel", "")).lower() == "predominante":
            linea_to_modos[lm["linea_id"]].add(lm["modo_id"])
    sub_to_modos = {}
    for s in sublineas:
        if s.get("linea") and s["linea"] in linea_to_modos:
            sub_to_modos[s["id"]] = linea_to_modos[s["linea"]]
    sub_ids_modo = sorted(sub_to_modos.keys())
    for i, a in enumerate(sub_ids_modo):
        for b in sub_ids_modo[i + 1:]:
            if sub_to_modos[a] & sub_to_modos[b]:
                edges["coincidencia_modo"].append({"source": a, "target": b})

    # (g) Proximidad semántica declarada
    for p in proximidad:
        edges["proximidad_semantica"].append({
            "source": p["sublinea_a_id"],
            "target": p["sublinea_b_id"],
            "weight": p.get("afinidad", 0),
        })

    return edges


# ---------------------------------------------------------------------------
# API pública
# ---------------------------------------------------------------------------

def load(xlsx_path):
    """Lee el xlsx y devuelve un dict con la misma forma que producía el
    loader JS: entities + relations + edges. Llamado desde build_doc.py."""
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(str(xlsx_path), data_only=True)

    tabs = {sn: _read_sheet(wb, sn) for sn in wb.sheetnames}

    # 1) Entidades
    lineas = [{
        "id": r["id"], "nombre": r["nombre"],
        "descripcion": r.get("descripción") or r.get("descripcion") or "",
        "estado": r.get("estado") or "",
    } for r in tabs.get("01_Lineas", []) if r.get("id")]

    areas = [{
        "id": r.get("código") or r.get("codigo"),
        "nombre": r["nombre"],
        "descripcion": r.get("descripción") or r.get("descripcion") or "",
    } for r in tabs.get("03_Areas", []) if r.get("código") or r.get("codigo")]

    modos = [{
        "id": r["id"], "nombre": r["nombre"],
        "descripcion": r.get("descripción") or r.get("descripcion") or "",
    } for r in tabs.get("04_Modos", []) if r.get("id")]

    salidas = [{
        "id": r["id"], "nombre": r["nombre"],
        "descripcion": r.get("descripción") or r.get("descripcion") or "",
    } for r in tabs.get("05_Salidas", []) if r.get("id")]

    laboratorios = [{
        "id": r["id"], "nombre": r["nombre"],
        "descripcion": r.get("descripción") or r.get("descripcion") or "",
    } for r in tabs.get("06_Laboratorios", []) if r.get("id")]

    investigadores = [{
        "id": r["id"], "nombre": r["nombre"],
        "area_principal": r.get("área_principal") or r.get("area_principal") or "",
        "perfil_url": r.get("perfil_casiopea") or r.get("perfil_url") or "",
    } for r in tabs.get("07_Investigadores", []) if r.get("id")]

    # 2) Mapas nombre→id
    linea_n2i = _build_name_to_id(lineas, "01_Lineas")
    area_n2i = _build_name_to_id(areas, "03_Areas")
    modo_n2i = _build_name_to_id(modos, "04_Modos")
    salida_n2i = _build_name_to_id(salidas, "05_Salidas")
    lab_n2i = _build_name_to_id(laboratorios, "06_Laboratorios")
    inv_n2i = _build_name_to_id(investigadores, "07_Investigadores")

    # 3) Sublíneas: las columnas `línea` y `área` son nombres
    sublineas = []
    for r in tabs.get("02_Sublineas", []):
        if not r.get("id"):
            continue
        sublineas.append({
            "id": r["id"],
            "nombre": r["nombre"],
            "linea": _resolve(r.get("línea"), linea_n2i,
                              f"02_Sublineas[{r['id']}].línea"),
            "area": _resolve(r.get("área"), area_n2i,
                             f"02_Sublineas[{r['id']}].área"),
            "notas": r.get("notas") or "",
        })
    sub_n2i = _build_name_to_id(sublineas, "02_Sublineas")

    # 4) Nueva 08_Temas: filas (sublínea, investigador, tema)
    temas = []
    sublinea_tema = []
    for idx, r in enumerate(tabs.get("08_Temas", [])):
        if not (r.get("sublínea") and r.get("investigador") and r.get("tema")):
            continue
        tema_id = f"TEM-{idx + 1:04d}"
        inv_id = _resolve(r.get("investigador"), inv_n2i,
                          f"08_Temas[{idx}].investigador")
        sub_id = _resolve(r.get("sublínea"), sub_n2i,
                          f"08_Temas[{idx}].sublínea")
        if not inv_id or not sub_id:
            continue
        temas.append({"id": tema_id, "investigador": inv_id, "texto": r["tema"]})
        sublinea_tema.append({"sublinea_id": sub_id, "tema_id": tema_id})

    # 5) Resto de relaciones
    lab_linea = []
    for i, r in enumerate(tabs.get("10_Lab_Linea", [])):
        lab_id = _resolve(r.get("laboratorio"), lab_n2i,
                          f"10_Lab_Linea[{i}].laboratorio")
        linea_id = _resolve(r.get("línea"), linea_n2i,
                            f"10_Lab_Linea[{i}].línea")
        if lab_id and linea_id:
            lab_linea.append({"lab_id": lab_id, "linea_id": linea_id})

    lab_salida = []
    for i, r in enumerate(tabs.get("11_Lab_Salida", [])):
        lab_id = _resolve(r.get("laboratorio"), lab_n2i,
                          f"11_Lab_Salida[{i}].laboratorio")
        salida_id = _resolve(r.get("salida"), salida_n2i,
                             f"11_Lab_Salida[{i}].salida")
        if lab_id and salida_id:
            lab_salida.append({"lab_id": lab_id, "salida_id": salida_id})

    inv_lab = []
    for i, r in enumerate(tabs.get("12_Investigador_Lab", [])):
        inv_id = _resolve(r.get("investigador"), inv_n2i,
                          f"12_Investigador_Lab[{i}].investigador")
        lab_id = _resolve(r.get("laboratorio"), lab_n2i,
                          f"12_Investigador_Lab[{i}].laboratorio")
        if inv_id and lab_id:
            inv_lab.append({"investigador_id": inv_id, "lab_id": lab_id})

    inv_modo = []
    for i, r in enumerate(tabs.get("13_Investigador_Modo", [])):
        inv_id = _resolve(r.get("investigador"), inv_n2i,
                          f"13_Investigador_Modo[{i}].investigador")
        modo_id = _resolve(r.get("modo"), modo_n2i,
                           f"13_Investigador_Modo[{i}].modo")
        if inv_id and modo_id:
            inv_modo.append({"investigador_id": inv_id, "modo_id": modo_id})

    linea_modo = []
    for i, r in enumerate(tabs.get("14_Linea_Modo", [])):
        linea_id = _resolve(r.get("línea"), linea_n2i,
                            f"14_Linea_Modo[{i}].línea")
        modo_id = _resolve(r.get("modo"), modo_n2i,
                           f"14_Linea_Modo[{i}].modo")
        if linea_id and modo_id:
            linea_modo.append({
                "linea_id": linea_id, "modo_id": modo_id,
                "nivel": r.get("nivel") or "",
            })

    proximidad = []
    for i, r in enumerate(tabs.get("18_Proximidad_Tematica", [])):
        if r.get("estado") == "DESCARTADO":
            continue
        a_id = _resolve(r.get("sublínea_a"), sub_n2i,
                        f"18_Proximidad_Tematica[{i}].sublínea_a")
        b_id = _resolve(r.get("sublínea_b"), sub_n2i,
                        f"18_Proximidad_Tematica[{i}].sublínea_b")
        if not (a_id and b_id):
            continue
        try:
            afin = float(r.get("afinidad") or 0)
        except (TypeError, ValueError):
            afin = 0.0
        proximidad.append({
            "sublinea_a_id": a_id,
            "sublinea_b_id": b_id,
            "afinidad": afin,
            "razonamiento": r.get("razonamiento") or "",
            "estado": r.get("estado") or "",
        })

    # 6) Sello
    sello = {"foco": "", "texto": ""}
    for r in tabs.get("17_Sello", []):
        foco = r.get("Foco") or r.get("foco") or ""
        if isinstance(foco, str) and "ELEGIDO" in foco:
            sello = {
                "foco": foco.replace(" (ELEGIDO)", ""),
                "texto": r.get("Texto") or r.get("texto") or "",
            }
            break

    edges = _compute_edges(
        sublineas=sublineas, lab_linea=lab_linea, linea_modo=linea_modo,
        proximidad=proximidad, sublinea_tema=sublinea_tema, temas=temas,
    )

    return {
        "version": "xlsx-direct",
        "sello": sello,
        "lineas": lineas,
        "sublineas": sublineas,
        "areas": areas,
        "modos": modos,
        "salidas": salidas,
        "laboratorios": laboratorios,
        "investigadores": investigadores,
        "temas": temas,
        "relations": {
            "sublinea_tema": sublinea_tema,
            "lab_linea": lab_linea,
            "lab_salida": lab_salida,
            "inv_lab": inv_lab,
            "inv_modo": inv_modo,
            "linea_modo": linea_modo,
            "proximidad": proximidad,
        },
        "edges": edges,
    }
